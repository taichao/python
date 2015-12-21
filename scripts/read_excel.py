# coding: utf-8
#读取execel使用(支持07)
from openpyxl import Workbook
#写入excel使用(支持07)
from openpyxl import load_workbook
import os

def writeexcel07(path):

    wb=Workbook()
    #sheet=wb.add_sheet("xlwt3数据测试表")
    sheet=wb.create_sheet(0,"xlwt3数据测试表")

    value = [["名称", "hadoop编程实战", "hbase编程实战", "lucene编程实战"], ["价格", "52.3", "45", "36"], ["出版社", "机械工业出版社", "人民邮电出版社", "华夏人民出版社"], ["中文版式", "中", "英", "英"]]
    #for i in range(0,4):
        #for j in range(0,len(value[i])):
            #sheet.write(i,j,value[i][j])

            #sheet.append(value[i])
    sheet.cell(row = 1,column= 2).value="温度"
    wb.save(path)
    print("写入数据成功！")


def read07excel(path):
    wb2=load_workbook(path)
    #print(wb2.get_sheet_names())
    ws=wb2.get_sheet_by_name("Sheet1")
    row=ws.max_row
    col=ws.max_column
    print("列数: ",col)
    print("行数: ",row)

    with open("/Users/zhangtaichao/news.txt",'a',encoding='utf-8') as f:
        for i  in range(1,row):
            line = ''
            for j in range(0,col):
                line = line + str( ws.rows[i][j].value )
                if j != col - 1:
                    line = line + chr(1)
            f.write(line)
            if i != row:
                f.write(os.linesep)


read07path="/Users/zhangtaichao/news.xlsx";
read07excel(read07path)
